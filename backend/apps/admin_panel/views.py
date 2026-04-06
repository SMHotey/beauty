from datetime import datetime, timedelta, date
from decimal import Decimal
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import ExtractHour
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.models import User
from rest_framework import viewsets, status, permissions, generics
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError, PermissionDenied

from apps.staff.models import Master, MasterService
from apps.appointments.models import Appointment, AppointmentService
from apps.clients.models import Client
from apps.services.models import Service
from apps.reviews.models import Review
from apps.admin_panel.serializers import (
    DashboardStatsSerializer,
    MasterCalendarSerializer,
    AdminAppointmentSerializer,
    SalesReportSerializer,
)


class AdminPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)


class DashboardStatsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]
    serializer_class = DashboardStatsSerializer

    def get(self, request, *args, **kwargs):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from and date_to:
            try:
                start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone())
                end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone())
            except (ValueError, TypeError):
                raise ValidationError({'detail': 'Неверный формат даты. Используйте ISO 8601.'})
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        appointments = Appointment.objects.filter(
            datetime_start__gte=start,
            datetime_start__lte=end,
            status__in=['completed', 'confirmed'],
        )

        revenue = appointments.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
        appointments_count = appointments.count()
        masters_count = Master.objects.filter(is_active=True).count()
        clients_count = Client.objects.count()
        avg_check = revenue / appointments_count if appointments_count > 0 else None

        top_services_qs = (
            Appointment.objects.filter(
                datetime_start__gte=start,
                datetime_start__lte=end,
                status='completed',
            )
            .prefetch_related('services__service')
            .values('services__service__name')
            .annotate(
                count=Count('services__service'),
                revenue=Sum('services__price_at_booking'),
            )
            .order_by('-count')[:10]
        )

        top_services = [
            {
                'name': item['services__service__name'],
                'count': item['count'],
                'revenue': str(item['revenue'] or 0),
            }
            for item in top_services_qs
            if item['services__service__name']
        ]

        master_load_qs = (
            Appointment.objects.filter(
                datetime_start__gte=start,
                datetime_start__lte=end,
                status='completed',
            )
            .values('master__id', 'master__user__first_name', 'master__user__last_name')
            .annotate(
                appointments_count=Count('id'),
                revenue=Sum('total_price'),
            )
            .order_by('-appointments_count')
        )

        master_load = [
            {
                'master_id': item['master__id'],
                'master_name': f"{item['master__user__first_name'] or ''} {item['master__user__last_name'] or ''}".strip() or 'Мастер',
                'appointments_count': item['appointments_count'],
                'revenue': str(item['revenue'] or 0),
            }
            for item in master_load_qs
        ]

        stats = {
            'revenue': str(revenue),
            'appointments_count': appointments_count,
            'masters_count': masters_count,
            'clients_count': clients_count,
            'avg_check': str(avg_check) if avg_check else None,
            'top_services': top_services,
            'master_load': master_load,
        }

        serializer = self.get_serializer(stats)
        return Response(serializer.data)


class AdminMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [AdminPermission]
    queryset = Master.objects.select_related('user').prefetch_related('master_services__service').all()
    lookup_field = 'id'

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        page = self.paginate_queryset(queryset)
        if page is not None:
            from apps.staff.serializers import MasterListSerializer
            serializer = MasterListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        from apps.staff.serializers import MasterListSerializer
        serializer = MasterListSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        from apps.staff.serializers import MasterSerializer
        serializer = MasterSerializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        phone = request.data.get('phone')
        if not phone:
            raise ValidationError({'phone': 'Поле phone обязательно.'})

        first_name = request.data.get('first_name', '')
        last_name = request.data.get('last_name', '')
        password = request.data.get('password', 'master123456')
        username = phone

        if User.objects.filter(username=username).exists():
            raise ValidationError({'phone': 'Пользователь с таким номером уже существует.'})

        with __import__('django.db', fromlist=['transaction']).transaction.atomic():
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name,
            )

            working_hours = request.data.get('working_hours', {})
            break_slots = request.data.get('break_slots', [])
            vacations = request.data.get('vacations', [])

            master = Master.objects.create(
                user=user,
                phone=phone,
                bio=request.data.get('bio', ''),
                is_active=request.data.get('is_active', True),
                working_hours=working_hours,
                break_slots=break_slots,
                vacations=vacations,
            )

        from apps.staff.serializers import MasterSerializer
        serializer = MasterSerializer(master)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        if 'phone' in request.data or 'first_name' in request.data or 'last_name' in request.data:
            user = instance.user
            if 'phone' in request.data:
                user.username = request.data['phone']
                instance.phone = request.data['phone']
            if 'first_name' in request.data:
                user.first_name = request.data['first_name']
            if 'last_name' in request.data:
                user.last_name = request.data['last_name']
            user.save()

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        from apps.staff.serializers import MasterSerializer
        read_serializer = MasterSerializer(instance)
        return Response(read_serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def get_serializer_class(self):
        if self.action == 'list':
            from apps.staff.serializers import MasterListSerializer
            return MasterListSerializer
        from apps.staff.serializers import MasterSerializer
        return MasterSerializer


class AdminMasterServicesView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def post(self, request, master_id, *args, **kwargs):
        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'detail': 'Мастер не найден.'})

        service_id = request.data.get('service_id')
        if not service_id:
            raise ValidationError({'service_id': 'Поле service_id обязательно.'})

        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            raise ValidationError({'service_id': 'Услуга не найдена.'})

        if MasterService.objects.filter(master=master, service=service).exists():
            raise ValidationError({'detail': 'Эта услуга уже назначена мастеру.'})

        custom_price = request.data.get('custom_price')
        custom_duration = request.data.get('custom_duration_minutes')
        is_enabled = request.data.get('is_enabled', True)

        master_service = MasterService.objects.create(
            master=master,
            service=service,
            custom_price=custom_price,
            custom_duration_minutes=custom_duration,
            is_enabled=is_enabled,
        )

        from apps.staff.serializers import MasterServiceSerializer
        serializer = MasterServiceSerializer(master_service)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def delete(self, request, master_id, *args, **kwargs):
        service_id = request.data.get('service_id')
        if not service_id:
            raise ValidationError({'service_id': 'Поле service_id обязательно.'})

        try:
            master_service = MasterService.objects.get(master_id=master_id, service_id=service_id)
        except MasterService.DoesNotExist:
            raise ValidationError({'detail': 'Связь мастера и услуги не найдена.'})

        master_service.delete()
        return Response({'detail': 'Услуга удалена у мастера.'}, status=status.HTTP_204_NO_CONTENT)

    def patch(self, request, master_id, *args, **kwargs):
        service_id = request.data.get('service_id')
        if not service_id:
            raise ValidationError({'service_id': 'Поле service_id обязательно.'})

        try:
            master_service = MasterService.objects.get(master_id=master_id, service_id=service_id)
        except MasterService.DoesNotExist:
            raise ValidationError({'detail': 'Связь мастера и услуги не найдена.'})

        if 'custom_price' in request.data:
            master_service.custom_price = request.data['custom_price']
        if 'custom_duration_minutes' in request.data:
            master_service.custom_duration_minutes = request.data['custom_duration_minutes']
        if 'is_enabled' in request.data:
            master_service.is_enabled = request.data['is_enabled']

        master_service.save()

        from apps.staff.serializers import MasterServiceSerializer
        serializer = MasterServiceSerializer(master_service)
        return Response(serializer.data)


class AdminCalendarView(generics.GenericAPIView):
    permission_classes = [AdminPermission]
    serializer_class = MasterCalendarSerializer

    def get(self, request, *args, **kwargs):
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        master_id = request.query_params.get('master_id')

        if not month or not year:
            now = timezone.now()
            month = now.month
            year = now.year

        try:
            month = int(month)
            year = int(year)
        except (ValueError, TypeError):
            raise ValidationError({'detail': 'Неверный формат month или year.'})

        if month < 1 or month > 12:
            raise ValidationError({'month': 'Месяц должен быть от 1 до 12.'})

        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        masters = Master.objects.select_related('user')
        if master_id:
            masters = masters.filter(id=master_id)

        result = []
        for master in masters:
            appointments = Appointment.objects.filter(
                master=master,
                datetime_start__date__gte=start_date,
                datetime_start__date__lte=end_date,
            ).prefetch_related('services__service').select_related('client__user').order_by('datetime_start')

            appt_list = []
            for appt in appointments:
                client_name = 'Гость'
                if appt.client:
                    if appt.client.user:
                        full = f"{appt.client.user.first_name} {appt.client.user.last_name}".strip()
                        client_name = full or appt.client.phone
                    else:
                        client_name = appt.client.phone

                appt_list.append({
                    'id': appt.id,
                    'client_id': appt.client.id if appt.client else None,
                    'client_name': client_name,
                    'client_phone': appt.client.phone if appt.client else None,
                    'master_id': master.id,
                    'datetime_start': appt.datetime_start.isoformat(),
                    'datetime_end': appt.datetime_end.isoformat(),
                    'status': appt.status,
                    'total_price': str(appt.total_price),
                    'services': [s.service.name for s in appt.services.all()],
                    'comment': appt.comment or '',
                })

            master_name = f"{master.user.first_name} {master.user.last_name}".strip() or master.user.username

            result.append({
                'master_id': master.id,
                'master_name': master_name,
                'date': start_date.isoformat(),
                'appointments': appt_list,
            })

        return Response(result)


class AdminAppointmentViewSet(viewsets.ModelViewSet):
    permission_classes = [AdminPermission]
    serializer_class = AdminAppointmentSerializer
    lookup_field = 'id'

    def get_queryset(self):
        qs = Appointment.objects.prefetch_related('services__service').select_related(
            'master__user', 'client__user'
        ).order_by('-datetime_start')

        status_filter = self.request.query_params.get('status')
        master_id = self.request.query_params.get('master_id')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if status_filter:
            qs = qs.filter(status=status_filter)
        if master_id:
            qs = qs.filter(master_id=master_id)
        if date_from:
            try:
                qs = qs.filter(datetime_start__gte=datetime.fromisoformat(date_from))
            except (ValueError, TypeError):
                pass
        if date_to:
            try:
                qs = qs.filter(datetime_start__lte=datetime.fromisoformat(date_to))
            except (ValueError, TypeError):
                pass

        return qs

    def create(self, request, *args, **kwargs):
        master_id = request.data.get('master_id')
        client_phone = request.data.get('client_phone')
        service_ids = request.data.get('service_ids', [])
        datetime_start = request.data.get('datetime_start')
        datetime_end = request.data.get('datetime_end')
        comment = request.data.get('comment', '')

        if not master_id or not service_ids or not datetime_start or not datetime_end:
            raise ValidationError({
                'detail': 'Поля master_id, service_ids, datetime_start, datetime_end обязательны.'
            })

        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'master_id': 'Мастер не найден.'})

        client = None
        if client_phone:
            client = Client.objects.filter(phone=client_phone).first()
            if not client:
                client = Client.objects.create(phone=client_phone)

        try:
            dt_start = datetime.fromisoformat(datetime_start)
            dt_end = datetime.fromisoformat(datetime_end)
        except (ValueError, TypeError):
            raise ValidationError({'detail': 'Неверный формат даты.'})

        with __import__('django.db', fromlist=['transaction']).transaction.atomic():
            appointment = Appointment.objects.create(
                client=client,
                master=master,
                datetime_start=dt_start,
                datetime_end=dt_end,
                status='confirmed',
                comment=comment,
            )

            from apps.staff.models import MasterService
            for sid in service_ids:
                ms = MasterService.objects.filter(master=master, service_id=sid, is_enabled=True).first()
                if ms:
                    price = ms.price
                    duration = ms.duration_minutes
                else:
                    svc = Service.objects.get(id=sid)
                    price = svc.base_price
                    duration = svc.base_duration_minutes

                Appointment.objects.prefetch_related('services')
                from apps.appointments.models import AppointmentService
                AppointmentService.objects.create(
                    appointment=appointment,
                    service_id=sid,
                    price_at_booking=price,
                    duration_at_booking=duration,
                )

            appointment.total_price = appointment.services.aggregate(
                total=Sum('price_at_booking')
            )['total'] or 0
            appointment.save()

        serializer = self.get_serializer(appointment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        allowed_fields = ['status', 'comment', 'datetime_start', 'datetime_end', 'master_id']

        if 'status' in request.data:
            valid_statuses = [choice[0] for choice in Appointment.STATUS_CHOICES]
            if request.data['status'] not in valid_statuses:
                raise ValidationError({'status': f'Допустимые статусы: {", ".join(valid_statuses)}'})
            instance.status = request.data['status']

        if 'comment' in request.data:
            instance.comment = request.data['comment']

        if 'datetime_start' in request.data:
            try:
                instance.datetime_start = datetime.fromisoformat(request.data['datetime_start'])
            except (ValueError, TypeError):
                raise ValidationError({'datetime_start': 'Неверный формат даты.'})

        if 'datetime_end' in request.data:
            try:
                instance.datetime_end = datetime.fromisoformat(request.data['datetime_end'])
            except (ValueError, TypeError):
                raise ValidationError({'datetime_end': 'Неверный формат даты.'})

        if 'master_id' in request.data:
            try:
                instance.master = Master.objects.get(id=request.data['master_id'])
            except Master.DoesNotExist:
                raise ValidationError({'master_id': 'Мастер не найден.'})

        instance.save()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class SalesReportView(generics.GenericAPIView):
    permission_classes = [AdminPermission]
    serializer_class = SalesReportSerializer

    def get(self, request, *args, **kwargs):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from and date_to:
            try:
                start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone())
                end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone())
            except (ValueError, TypeError):
                raise ValidationError({'detail': 'Неверный формат даты.'})
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        daily_stats = (
            Appointment.objects.filter(
                datetime_start__gte=start,
                datetime_start__lte=end,
                status='completed',
            )
            .extra({'date': "DATE(datetime_start)"})
            .values('date')
            .annotate(
                revenue=Sum('total_price'),
                appointments_count=Count('id'),
                avg_check=Avg('total_price'),
            )
            .order_by('date')
        )

        report_data = []
        for item in daily_stats:
            report_data.append({
                'date': item['date'],
                'revenue': str(item['revenue'] or 0),
                'appointments_count': item['appointments_count'],
                'avg_check': str(item['avg_check']) if item['avg_check'] else None,
            })

        export_format = request.query_params.get('format', 'json')

        if export_format == 'xlsx':
            return self._export_xlsx(report_data, start, end)

        return Response(report_data)

    def _export_xlsx(self, data, start, end):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Отчёт по продажам'

        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f'Отчёт по продажам: {start.strftime("%d.%m.%Y")} — {end.strftime("%d.%m.%Y")}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Дата', 'Выручка (₽)', 'Кол-во записей', 'Средний чек (₽)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=item['date']).border = thin_border
            revenue_cell = ws.cell(row=row_idx, column=2, value=float(item['revenue']))
            revenue_cell.number_format = '#,##0.00'
            revenue_cell.border = thin_border
            ws.cell(row=row_idx, column=3, value=item['appointments_count']).border = thin_border

            avg_check = item['avg_check']
            if avg_check:
                avg_cell = ws.cell(row=row_idx, column=4, value=float(avg_check))
                avg_cell.number_format = '#,##0.00'
            else:
                avg_cell = ws.cell(row=row_idx, column=4, value=0)
            avg_cell.border = thin_border

        total_revenue = sum(float(item['revenue']) for item in data)
        total_appointments = sum(item['appointments_count'] for item in data)
        total_row = len(data) + 4

        ws.cell(row=total_row, column=1, value='ИТОГО').font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = thin_border
        rev_cell = ws.cell(row=total_row, column=2, value=total_revenue)
        rev_cell.number_format = '#,##0.00'
        rev_cell.font = Font(bold=True)
        rev_cell.border = thin_border
        ws.cell(row=total_row, column=3, value=total_appointments).font = Font(bold=True)
        ws.cell(row=total_row, column=3).border = thin_border
        avg_total = total_revenue / total_appointments if total_appointments > 0 else 0
        avg_cell = ws.cell(row=total_row, column=4, value=avg_total)
        avg_cell.number_format = '#,##0.00'
        avg_cell.font = Font(bold=True)
        avg_cell.border = thin_border

        for col_idx in range(1, 5):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, total_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = max_length + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start.strftime("%Y%m%d")}_{end.strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


class SmsBroadcastView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def post(self, request, *args, **kwargs):
        message = request.data.get('message')
        if not message:
            raise ValidationError({'message': 'Поле message обязательно.'})

        client_ids = request.data.get('client_ids')
        send_to_all = request.data.get('send_to_all', False)

        if send_to_all:
            phones = list(Client.objects.values_list('phone', flat=True))
        elif client_ids:
            phones = list(Client.objects.filter(id__in=client_ids).values_list('phone', flat=True))
        else:
            raise ValidationError({'detail': 'Укажите client_ids или send_to_all=true.'})

        if not phones:
            return Response({'detail': 'Нет клиентов для рассылки.'}, status=status.HTTP_400_BAD_REQUEST)

        sent_count = len(phones)

        try:
            from apps.admin_panel.tasks import send_bulk_sms
            send_bulk_sms.delay(phones, message)
            task_status = 'queued'
        except (ImportError, Exception):
            # Fallback when Celery/Redis is unavailable (dev/testing)
            from apps.auth_app.sms_service import SMSService
            for phone in phones:
                SMSService.send_code(phone, code='BROADCAST')
            task_status = 'sent_stub'

        return Response({
            'detail': f'Рассылка запущена для {sent_count} клиентов.',
            'phones_count': sent_count,
            'task_status': task_status,
        }, status=status.HTTP_200_OK)


class AdminServiceViewSet(viewsets.ModelViewSet):
    permission_classes = [AdminPermission]
    queryset = Service.objects.select_related('category').all()
    lookup_field = 'id'

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            from apps.services.serializers import ServiceSerializer
            serializer = ServiceSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        from apps.services.serializers import ServiceSerializer
        serializer = ServiceSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        from apps.services.serializers import ServiceSerializer
        serializer = ServiceSerializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        from apps.services.serializers import ServiceSerializer
        serializer = ServiceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        from apps.services.serializers import ServiceSerializer
        serializer = ServiceSerializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs, partial=True)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterStatsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, master_id, *args, **kwargs):
        try:
            master = Master.objects.select_related('user').get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'detail': 'Мастер не найден.'})

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from and date_to:
            try:
                start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone())
                end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone())
            except (ValueError, TypeError):
                raise ValidationError({'detail': 'Неверный формат даты.'})
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        completed = Appointment.objects.filter(
            master=master,
            status='completed',
            datetime_start__gte=start,
            datetime_start__lte=end,
        )

        total_revenue = completed.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
        total_appointments = completed.count()
        total_hours = completed.aggregate(
            hours=Sum(
                (models.F('datetime_end') - models.F('datetime_start')) / timedelta(hours=1)
            )
        )['hours'] or 0

        service_stats_qs = (
            Appointment.objects.filter(
                master=master,
                status='completed',
                datetime_start__gte=start,
                datetime_start__lte=end,
            )
            .prefetch_related('services__service')
            .values('services__service__id', 'services__service__name')
            .annotate(
                count=Count('services__service'),
                revenue=Sum('services__price_at_booking'),
            )
            .order_by('-count')
        )

        service_stats = [
            {
                'service_id': item['services__service__id'],
                'service_name': item['services__service__name'],
                'count': item['count'],
                'revenue': str(item['revenue'] or 0),
            }
            for item in service_stats_qs
            if item['services__service__name']
        ]

        reviews_qs = Review.objects.filter(
            appointment__master=master,
        ).select_related('client__user').order_by('-created_at')[:20]

        reviews = [
            {
                'id': r.id,
                'client_name': f"{r.client.user.first_name} {r.client.user.last_name}".strip() if r.client and r.client.user else r.client.phone if r.client else 'Гость',
                'rating': r.rating,
                'comment': r.comment,
                'created_at': r.created_at.isoformat(),
            }
            for r in reviews_qs
        ]

        avg_rating = Review.objects.filter(
            appointment__master=master,
        ).aggregate(avg=Avg('rating'))['avg']

        master_name = f"{master.user.first_name} {master.user.last_name}".strip() or master.user.username

        return Response({
            'master_id': master.id,
            'master_name': master_name,
            'bio': master.bio,
            'phone': master.phone,
            'is_active': master.is_active,
            'total_revenue': str(total_revenue),
            'total_appointments': total_appointments,
            'total_hours': round(total_hours, 1),
            'avg_rating': round(avg_rating, 1) if avg_rating else None,
            'service_stats': service_stats,
            'recent_reviews': reviews,
        })


class ClientStatsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, client_id, *args, **kwargs):
        try:
            client = Client.objects.select_related('user').get(id=client_id)
        except Client.DoesNotExist:
            raise ValidationError({'detail': 'Клиент не найден.'})

        user = client.user
        client_name = f"{user.first_name} {user.last_name}".strip() if user else client.phone

        appointments = Appointment.objects.filter(client=client).order_by('-datetime_start')
        total_visits = appointments.count()
        total_spent = appointments.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
        completed = appointments.filter(status='completed')

        service_qs = (
            completed
            .prefetch_related('services__service')
            .values('services__service__id', 'services__service__name')
            .annotate(
                count=Count('services__service'),
                spent=Sum('services__price_at_booking'),
            )
            .order_by('-count')
        )

        service_stats = [
            {
                'service_id': item['services__service__id'],
                'service_name': item['services__service__name'],
                'count': item['count'],
                'spent': str(item['spent'] or 0),
            }
            for item in service_qs
            if item['services__service__name']
        ]

        recent_appointments = [
            {
                'id': a.id,
                'master_name': f"{a.master.user.first_name} {a.master.user.last_name}".strip() if a.master and a.master.user else '—',
                'datetime_start': a.datetime_start.isoformat(),
                'datetime_end': a.datetime_end.isoformat() if a.datetime_end else '',
                'status': a.status,
                'total_price': str(a.total_price or 0),
                'services': [
                    {'name': s.service.name, 'price': str(s.price_at_booking)}
                    for s in a.services.all()
                ],
            }
            for a in appointments[:20]
        ]

        return Response({
            'client_id': client.id,
            'client_name': client_name,
            'phone': client.phone,
            'email': user.email if user else '',
            'bonus_balance': client.bonus_balance,
            'referral_code': client.referral_code,
            'referred_by': client.referred_by.phone if client.referred_by else None,
            'total_visits': total_visits,
            'total_spent': str(total_spent),
            'completed_visits': completed.count(),
            'service_stats': service_stats,
            'recent_appointments': recent_appointments,
        })


class ServiceStatsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, service_id, *args, **kwargs):
        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            raise ValidationError({'detail': 'Услуга не найдена.'})

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from and date_to:
            try:
                start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone())
                end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone())
                end = end.replace(hour=23, minute=59, second=59)
            except (ValueError, TypeError):
                raise ValidationError({'detail': 'Неверный формат даты.'})
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        masters = Master.objects.select_related('user').filter(
            appointments__services__service=service,
            appointments__datetime_start__gte=start,
            appointments__datetime_start__lte=end,
        ).distinct()

        result = []
        for master in masters:
            master_name = f"{master.user.first_name} {master.user.last_name}".strip() or master.user.username

            service_qs = Appointment.objects.filter(
                master=master,
                services__service=service,
                datetime_start__gte=start,
                datetime_start__lte=end,
            )

            total_count = service_qs.count()
            total_revenue = service_qs.aggregate(total=Sum('services__price_at_booking'))['total'] or Decimal('0')

            result.append({
                'master_id': master.id,
                'master_name': master_name,
                'services': [
                    {
                        'service_id': service.id,
                        'service_name': service.name,
                        'count': total_count,
                        'revenue': str(total_revenue),
                    }
                ],
                'total_count': total_count,
                'total_revenue': str(total_revenue),
            })

        return Response(result)


class MasterServiceStatsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, master_id, service_id, *args, **kwargs):
        try:
            master = Master.objects.select_related('user').get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'detail': 'Мастер не найден.'})

        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            raise ValidationError({'detail': 'Услуга не найдена.'})

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if date_from and date_to:
            try:
                start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone())
                end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone())
                end = end.replace(hour=23, minute=59, second=59)
            except (ValueError, TypeError):
                raise ValidationError({'detail': 'Неверный формат даты.'})
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        completed = Appointment.objects.filter(
            master=master,
            status='completed',
            services__service=service,
            datetime_start__gte=start,
            datetime_start__lte=end,
        )

        total_count = completed.count()
        total_revenue = completed.aggregate(total=Sum('services__price_at_booking'))['total'] or Decimal('0')

        master_name = f"{master.user.first_name} {master.user.last_name}".strip() or master.user.username

        return Response({
            'master_id': master.id,
            'master_name': master_name,
            'service_id': service.id,
            'service_name': service.name,
            'total_count': total_count,
            'total_revenue': str(total_revenue),
            'date_from': start.isoformat(),
            'date_to': end.isoformat(),
        })


class MasterPermissionsView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, master_id, *args, **kwargs):
        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'detail': 'Мастер не найден.'})

        perms, _ = MasterPermission.objects.get_or_create(master=master)
        return Response({
            'master_id': master.id,
            'master_name': f"{master.user.first_name} {master.user.last_name}".strip() or master.user.username,
            'can_edit_schedule': perms.can_edit_schedule,
            'can_reply_reviews': perms.can_reply_reviews,
        })

    def patch(self, request, master_id, *args, **kwargs):
        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            raise ValidationError({'detail': 'Мастер не найден.'})

        perms, _ = MasterPermission.objects.get_or_create(master=master)
        if 'can_edit_schedule' in request.data:
            perms.can_edit_schedule = request.data['can_edit_schedule']
        if 'can_reply_reviews' in request.data:
            perms.can_reply_reviews = request.data['can_reply_reviews']
        perms.save()

        return Response({
            'master_id': master.id,
            'can_edit_schedule': perms.can_edit_schedule,
            'can_reply_reviews': perms.can_reply_reviews,
        })


class AdminStatsExportView(generics.GenericAPIView):
    permission_classes = [AdminPermission]

    def get(self, request, *args, **kwargs):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        master_id = request.query_params.get('master_id')
        service_id = request.query_params.get('service_id')
        payment_method = request.query_params.get('payment_method')
        export_format = request.query_params.get('format', 'json')

        qs = Appointment.objects.filter(status='completed').order_by('-datetime_start')

        if date_from:
            qs = qs.filter(datetime_start__gte=datetime.fromisoformat(date_from).replace(tzinfo=timezone.get_current_timezone()))
        if date_to:
            qs = qs.filter(datetime_start__lte=datetime.fromisoformat(date_to).replace(tzinfo=timezone.get_current_timezone(), hour=23, minute=59, second=59))
        if master_id:
            qs = qs.filter(master_id=master_id)
        if service_id:
            qs = qs.filter(services__service_id=service_id)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)

        if export_format == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Статистика'

            headers = ['Дата', 'Время', 'Мастер', 'Клиент', 'Услуги', 'Способ оплаты', 'Сумма']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

            for row_idx, apt in enumerate(qs, 2):
                client_name = 'Гость'
                if apt.client:
                    client_name = f"{apt.client.user.first_name} {apt.client.user.last_name}".strip() or apt.client.user.username
                master_name = f"{apt.master.user.first_name} {apt.master.user.last_name}".strip() if apt.master.user else ''
                services = ', '.join([s.service.name for s in apt.services.all()])
                payment_labels = dict(Appointment.PAYMENT_CHOICES)

                ws.cell(row=row_idx, column=1, value=apt.datetime_start.strftime('%d.%m.%Y'))
                ws.cell(row=row_idx, column=2, value=apt.datetime_start.strftime('%H:%M'))
                ws.cell(row=row_idx, column=3, value=master_name)
                ws.cell(row=row_idx, column=4, value=client_name)
                ws.cell(row=row_idx, column=5, value=services)
                ws.cell(row=row_idx, column=6, value=payment_labels.get(apt.payment_method, apt.payment_method))
                ws.cell(row=row_idx, column=7, value=float(apt.total_price or 0))

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=stats.xlsx'
            return response

        data = []
        for apt in qs:
            client_name = 'Гость'
            if apt.client:
                client_name = f"{apt.client.user.first_name} {apt.client.user.last_name}".strip() or apt.client.user.username
            master_name = f"{apt.master.user.first_name} {apt.master.user.last_name}".strip() if apt.master.user else ''
            services = [s.service.name for s in apt.services.all()]

            data.append({
                'datetime': apt.datetime_start.isoformat(),
                'master': master_name,
                'client': client_name,
                'services': services,
                'payment_method': apt.payment_method,
                'total_price': str(apt.total_price or 0),
            })

        total_revenue = qs.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
        return Response({
            'appointments': data,
            'total_count': qs.count(),
            'total_revenue': str(total_revenue),
        })
